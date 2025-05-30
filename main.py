from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, ContextTypes, JobQueue
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from telegram import Bot
import asyncio
from dotenv import load_dotenv
import os

load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")

GROUP_IDS = [-1234567878, -4123456789, -1234567890, -12345678] #Ganti Dengan Grub ID
file_name = "user_data/user_data.xlsx"
HARGA_PER_BULAN = 15.113   # Harga Per satu bulan
ADMIN_CHAT_IDS = ["1234567890", "1234567890"]  # Ganti dengan chat_id admin
bot = Bot(token=(TOKEN))

# Fungsi untuk membaca dan memperbarui file Excel
def update_excel_file(file_name):
    # Membuka file xlsx
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    
    # Menyimpan tanggal hari ini untuk perhitungan
    today = datetime.now()
    
    # Periksa apakah file kosong (tidak ada data selain header)
    if sheet.max_row < 2:
        print("File Excel kosong atau tidak memiliki data yang cukup.")
        wb.close()
        return

    data_updated = False  # Flag untuk mengecek apakah ada data yang diperbarui

    # Iterasi baris-baris data
    for row in range(2, sheet.max_row + 1):  # Mulai dari baris kedua
        # Ambil data Sisa Waktu dan Tanggal Kadaluarsa
        sisa_waktu = sheet.cell(row=row, column=6).value  # Kolom 6 untuk Sisa Waktu
        tanggal_kadaluarsa = sheet.cell(row=row, column=5).value  # Kolom 5 untuk Tanggal Kadaluarsa
        
        # Jika sel kosong, lewati baris ini
        if not sisa_waktu or not tanggal_kadaluarsa:
            print(f"Data kosong di baris {row}, dilewati.")
            continue

        # Pastikan Tanggal Kadaluarsa adalah datetime, jika berupa string konversikan
        if isinstance(tanggal_kadaluarsa, str):
            try:
                tanggal_kadaluarsa = datetime.strptime(tanggal_kadaluarsa, '%d %B %Y')  # Ganti format sesuai format di Excel
            except ValueError:
                print(f"Format tanggal tidak valid di baris {row}: {tanggal_kadaluarsa}")
                continue  # Lewati baris ini jika format tidak valid

        # Mengurangi Sisa Waktu
        if isinstance(sisa_waktu, str) and "hari" in sisa_waktu:
            try:
                days_left = int(sisa_waktu.split()[0])  # Ambil angka hari dari Sisa Waktu
            except ValueError:
                print(f"Format Sisa Waktu tidak valid di baris {row}: {sisa_waktu}")
                continue

            new_days_left = days_left - 1  # Kurangi 1 hari
            sheet.cell(row=row, column=6, value=f"{new_days_left} hari")  # Update Sisa Waktu

            # Memperbarui Tanggal Kadaluarsa
            new_expiration_date = tanggal_kadaluarsa - timedelta(days=1)  # Kurangi 1 hari
            sheet.cell(row=row, column=5, value=new_expiration_date.strftime('%d %B %Y'))  # Update Tanggal Kadaluarsa

            data_updated = True  # Set flag menjadi True jika data diperbarui

    # Simpan perubahan hanya jika ada data yang diperbarui
    if data_updated:
        wb.save(file_name)
        print("File berhasil diperbarui.")
    else:
        print("Tidak ada data yang diperbarui.")

    wb.close()


    # Simpan perubahan
    wb.save(file_name)

# Fungsi untuk memulai bot
async def start(update: Update, context: CallbackContext):
    user_name = update.message.from_user.first_name
    keyboard = [
        [InlineKeyboardButton("💎 𝐁𝐄𝐋𝐈 𝐕𝐕𝐈𝐏 💎", callback_data="beli_vvip")],
        [
            InlineKeyboardButton("🆘 𝐁𝐀𝐍𝐓𝐔𝐀𝐍? 🆘", url="https://t.me/Userr0998"),
            InlineKeyboardButton("📊 𝐒𝐓𝐀𝐓𝐔𝐒", callback_data="status")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(f"""
ꜱᴇʟᴀᴍᴀᴛ ᴅᴀᴛᴀɴɢ {user_name}.

ꜱᴇɴᴀɴɢ ʙᴇʀᴛᴇᴍᴜ ᴅᴇɴɢᴀɴᴍᴜ ᴅɪ ᴘᴀʏᴍᴇɴᴛ ᴠᴠɪᴘ..
ʟᴀᴋᴜᴋᴀɴ ᴘᴇᴍʙᴀʏᴀʀᴀɴ ᴜɴᴛᴜᴋ ᴍᴀꜱᴜᴋ ᴋᴇ ᴄʜᴀɴɴᴇʟ ᴠᴠɪᴘ

ᴘᴇʀɪɴɢᴀᴛᴀɴ : ʟᴀᴋᴜᴋᴀɴ ᴘᴇᴍʙᴀʏᴀʀᴀɴ ᴅᴇɴɢᴀɴ Qʀɪꜱ ʏᴀɴɢ ꜱᴜᴅᴀʜ ᴅɪꜱᴇᴅɪᴀᴋᴀɴ ᴀᴅᴍɪɴ, ᴊᴀɴɢᴀɴ Qʀɪꜱ ʟᴀɪɴ❗❗
""", reply_markup=reply_markup)

async def beli_vvip(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    context.user_data['bulan'] = 1
    context.user_data['harga'] = HARGA_PER_BULAN
    context.user_data['tanggal_akhir'] = (datetime.now() + timedelta(days=30)).strftime('%d %B %Y')

    await send_keranjang(update, context)

async def send_keranjang(update: Update, context: CallbackContext):
    query = update.callback_query
    bulan = context.user_data['bulan']
    harga = context.user_data['harga']
    tanggal_akhir = context.user_data['tanggal_akhir']

    # Menyimpan tampilan sebelumnya
    context.user_data['previous_view'] = "keranjang"

    keyboard = [
        [InlineKeyboardButton("-1", callback_data="kurang_bulan"), InlineKeyboardButton("+1", callback_data="tambah_bulan")],
        [InlineKeyboardButton("KEMBALI", callback_data="batal"), InlineKeyboardButton("LANJUT", callback_data="lanjut")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        text=f"𝙋𝙀𝙈𝘽𝙀𝙇𝙄𝘼𝙉 :\nVVIP : {bulan} Bulan\n\nVVIP BERAKHIR : {tanggal_akhir}\n\nTOTAL HARGA : Rp. {harga}",
        reply_markup=reply_markup
    )

async def adjust_month(update: Update, context: CallbackContext, change: int):
    query = update.callback_query
    bulan = context.user_data.get('bulan', 0) + change

    if bulan < 1:
        bulan = 1
    elif bulan > 12:  # Misalnya bulan tidak bisa lebih dari 12
        bulan = 12

    context.user_data['bulan'] = bulan
    context.user_data['harga'] = bulan * HARGA_PER_BULAN
    context.user_data['tanggal_akhir'] = (datetime.now() + timedelta(days=30*bulan)).strftime('%d %B %Y')
    await send_keranjang(update, context)

async def lanjut(update: Update, context: CallbackContext):
    query = update.callback_query
    user_name = query.from_user.first_name
    # bulan = context.user_data['bulan']
    # harga = context.user_data['harga']
    bulan = context.user_data.get('bulan', 1)  # Default ke 1 jika tidak ada
    harga = context.user_data.get('harga', bulan * 15.113)
    user_id = query.from_user.id
    chat_id = update.effective_chat.id
    

    keyboard = [
        [InlineKeyboardButton("INFO PEMBAYARAN", url="https://t.me/Ex_Storee")],
        [InlineKeyboardButton("BATALKAN", callback_data="batal")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        text=f"𝙄𝙉𝙁𝙊𝙍𝙈𝘼𝙎𝙄 𝙋𝙀𝙈𝘽𝘼𝙔𝘼𝙍𝘼𝙉:\n\nID: {query.from_user.id}\nNAMA: {user_name}\nBULAN: {bulan} Bulan\nHARGA: Rp. {harga}\n\nSilahkan Klik Tombol INFO PEMBAYARAN di bawah ini untuk metode pembayarannya. Jika sudah membayar, silahkan kirim screenshot bukti pembayaran di sini.",
        reply_markup=reply_markup
    )
    
    # Set Waktu berakhir expired pembayaran selama 5 Menit
    context.job_queue.run_once(pembayaran_expired, when=300, data={"chat_id": chat_id})

async def pembayaran_expired(context: CallbackContext):
    chat_id = context.job.data.get("chat_id")

    if context.job.data.get("expired") is not None:
        context.job.data["expired"] = True  # Jika expired sudah ada, perbarui status
    else:
        # Jika tidak, gunakan pesan default
        await context.bot.send_message(chat_id=chat_id, text="Waktu pembayaran telah habis. Silakan coba lagi.")
        context.job.data["expired"] = True

async def batal(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    # Mengambil nama pengguna dari callback_query
    user_name = query.from_user.first_name
    
    # Menyiapkan keyboard untuk kembali ke tampilan awal
    keyboard = [
        [InlineKeyboardButton("💎 𝐁𝐄𝐋𝐈 𝐕𝐕𝐈𝐏 💎", callback_data="beli_vvip")],
        [
            InlineKeyboardButton("🆘 𝐁𝐀𝐍𝐓𝐔𝐀𝐍? 🆘", url="https://t.me/Userr0998"),
            InlineKeyboardButton("📊 𝐒𝐓𝐀𝐓𝐔𝐒", callback_data="status")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Mengirimkan pesan baru dengan tombol kembali ke start
    await query.edit_message_text(f"""
ꜱᴇʟᴀᴍᴀᴛ ᴅᴀᴛᴀɴɢ {user_name}.

ꜱᴇɴᴀɴɢ ʙᴇʀᴛᴇᴍᴜ ᴅᴇɴɢᴀɴᴍᴜ ᴅɪ ᴘᴀʏᴍᴇɴᴛ ᴠᴠɪᴘ..
ʟᴀᴋᴜᴋᴀɴ ᴘᴇᴍʙᴀʏᴀʀᴀɴ ᴜɴᴛᴜᴋ ᴍᴀꜱᴜᴋ ᴋᴇ ᴄʜᴀɴɴᴇʟ ᴠᴠɪᴘ

ᴘᴇʀɪɴɢᴀᴛᴀɴ : ʟᴀᴋᴜᴋᴀɴ ᴘᴇᴍʙᴀʏᴀʀᴀɴ ᴅᴇɴɢᴀɴ Qʀɪꜱ ʏᴀɴɢ ꜱᴜᴅᴀʜ ᴅɪꜱᴇᴅɪᴀᴋᴀɴ ᴀᴅᴍɪɴ, ᴊᴀɴɢᴀɴ Qʀɪꜱ ʟᴀɪɴ❗❗""", reply_markup=reply_markup)


async def handle_photo(update, context):
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    user_name = update.message.from_user.first_name
    username = update.message.from_user.username
    photo_file = update.message.photo[-1].file_id

    bulan = context.user_data.get('bulan', 1)
    harga = context.user_data.get('harga', bulan * HARGA_PER_BULAN)

    # Kirimkan foto ke admin dengan pesan yang disesuaikan dan inline buttons
    message = f"Informasi Pembayaran dari:\n\nID: {user_id}\nNama: {user_name}\nUsername : {username}\nBulan : {bulan} bulan\nTotal Harga : {harga}\n\nKlik Yes jika pembayaran benar dan Anda akan menambahkan {user_name} ke dalam grup, klik No jika pembayaran tidak sah/pembayaran palsu."

    # Membuat inline keyboard dengan dua tombol: Yes dan No
    keyboard = [
        [InlineKeyboardButton("YES", callback_data=f"yes_{user_id}")],
        [InlineKeyboardButton("NO", callback_data=f"no_{user_id}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Mengirim foto bersama pesan kepada admin
    for admin_id in ADMIN_CHAT_IDS:
        try:
        # Cek apakah foto valid
            if not photo_file:
                raise ValueError("Foto tidak ditemukan atau tidak valid.")
        
        # Cek apakah admin sudah memulai percakapan
            chat_member = await context.bot.get_chat_member(admin_id, admin_id)
            if chat_member.status not in ["member", "administrator", "creator"]:
                print(f"Admin {admin_id} belum memulai percakapan dengan bot.")
                continue
        
        # Kirim foto ke admin
            await context.bot.send_photo(
                chat_id=admin_id,
                photo=photo_file,
                caption=message,
                reply_markup=reply_markup
            )
            print(f"Pesan berhasil dikirim ke admin {admin_id}")

        except Exception as e:
            if "bot can't initiate conversation" in str(e):
                print(f"Gagal mengirim pesan ke admin {admin_id}: Admin belum memulai chat dengan bot.")
            else:
                print(f"Gagal mengirim pesan ke admin {admin_id}: {e}")

    await context.bot.send_message(
                            chat_id=user_id,
                            text=f"Foto pembayaran sedang diproses oleh admin silahkan menunggu maksimal 2 Jam"
                        )

async def button(update, context):
    query = update.callback_query 
    # action, user_id = query.data.split('_')
    # user_id = int(user_id)
    action = query.data
    user_name = query.from_user.first_name
    user_id = query.from_user.id
        

    chat_data = context.chat_data

    if "_" in action:
        action, user_id = action.split('_')
        user_id = int(user_id)
    else:
        user_id = query.from_user.id

    # Default nilai
    bulan = context.user_data.get('bulan', 1)  # Ambil jumlah bulan yang dipilih user
    harga = context.user_data.get('harga', bulan * 15.113)
    tanggal_akhir = (datetime.now() + timedelta(days=30 * bulan)).strftime('%d %B %Y')

    # Jika tombol "STATUS" ditekan
    if action == "status":
        try:
            df = pd.read_excel(file_name)
            user_data = df[df['Chat ID'] == user_id]

            if not user_data.empty:
                # Jika data ditemukan, ambil status dan tanggal kadaluarsa
                status_vip = user_data['Status VIP'].values[0]
                tanggal_akhir = user_data['Tanggal Kadaluarsa'].values[0]
                sisa_hari = (datetime.strptime(tanggal_akhir, "%d %B %Y") - datetime.now()).days
            else:
                # Jika data tidak ditemukan, set status dan tanggal kadaluarsa ke default
                status_vip = "Tidak Aktif"
                tanggal_akhir = "-"
                sisa_hari = 0
            
            # Membuat inline keyboard dengan tombol "KEMBALI"
            keyboard = [[InlineKeyboardButton("KEMBALI", callback_data="batal")]]
            reply_markup = InlineKeyboardMarkup(keyboard)

            # Kirim pesan dengan status VIP dan informasi lainnya
            await query.edit_message_text(
                text=f"""
𝐒𝐓𝐀𝐓𝐔𝐒 𝐕𝐕𝐈𝐏 :

🆔 Chat ID: {user_id}
👤 Nama Lengkap: {query.from_user.first_name}
👤 Username: @{query.from_user.username or 'Tidak Ada'}
📅 Status VIP: {status_vip}
📆 Tanggal Kadaluarsa: {tanggal_akhir or '-'}
⏳ Sisa Waktu: {sisa_hari} hari
                """,
                reply_markup=reply_markup
            )
        except FileNotFoundError:
            await query.answer("Data tidak ditemukan.")

        return
        
    admin_messages = {}
    original_message_id = query.message.message_id
    action = query.data.split('_')[0]  # Mendapatkan action (yes/no)
    user_id_from_data = int(query.data.split('_')[1])
    
    await query.message.edit_reply_markup(reply_markup=None)
# Jika admin menekan Yes
    if action == "yes":
        try:
            try:
                existing_data = pd.read_excel(file_name)
            except FileNotFoundError:
                existing_data = pd.DataFrame(columns=['Chat ID', 'Nama Lengkap', 'Username', 'Status VIP', 'Tanggal Kadaluarsa', 'Sisa Waktu'])

            current_date = datetime.now()
            
            # Get user data directly from the callback data
            user_message = query.message
            user_info = user_message.caption.split('\n') if user_message.caption else []
            
            # Extract user information from the message caption
            user_data = {}
            for line in user_info:
                if "ID:" in line:
                    user_data['id'] = int(line.split(': ')[1])
                elif "Nama:" in line:
                    user_data['name'] = line.split(': ')[1]
                elif "Username :" in line:
                    user_data['username'] = line.split(': ')[1]
                elif "Bulan :" in line:
                    user_data['bulan'] = int(line.split(': ')[1].split()[0])
            
            # Use extracted data
            payment_user_id = user_data['id']
            bulan = user_data.get('bulan', 1)  # Default to 1 if not found
            
            # Cek apakah user sudah ada di database
            user_exists = existing_data['Chat ID'] == payment_user_id
    
            if any(user_exists):
                # Update data user yang sudah ada
                user_index = user_exists.idxmax()
                current_expiry = datetime.strptime(existing_data.loc[user_index, 'Tanggal Kadaluarsa'], '%d %B %Y')
        
                # Jika masih aktif, tambahkan ke tanggal existing
                if current_expiry > current_date:
                    new_expiry = current_expiry + timedelta(days=30 * bulan)
                else:
                    # Jika sudah expired, mulai dari sekarang
                    new_expiry = current_date + timedelta(days=30 * bulan)
        
                existing_data.loc[user_index, 'Tanggal Kadaluarsa'] = new_expiry.strftime('%d %B %Y')
                existing_data.loc[user_index, 'Sisa Waktu'] = f"{(new_expiry - current_date).days} hari"
                existing_data.loc[user_index, 'Status VIP'] = "Aktif"
            else:
                # Tambah user baru
                tanggal_akhir = (current_date + timedelta(days=30 * bulan)).strftime('%d %B %Y')
                sisa_hari = 30 * bulan
                
                new_user = {
                    'Chat ID': payment_user_id,
                    'Nama Lengkap': user_data['name'],
                    'Username': user_data['username'],
                    'Status VIP': "Aktif",
                    'Tanggal Kadaluarsa': tanggal_akhir,
                    'Sisa Waktu': f"{sisa_hari} hari"
                }
                existing_data = pd.concat([existing_data, pd.DataFrame([new_user])], ignore_index=True)

            # Simpan ke Excel
            existing_data.to_excel(file_name, index=False)
        # Kirim pesan sukses dan link grup
            keyboard = [
                [
                    InlineKeyboardButton("CHANNEL 1", url="https://"),
                    InlineKeyboardButton("CHANNEL 2", url="https://"),
                ],
                [
                    InlineKeyboardButton("GRUB 1", url="https://"),
                    InlineKeyboardButton("GRUB 2", url="https://"),
                ]
            ]

            reply_markup = InlineKeyboardMarkup(keyboard)
            await context.bot.send_photo(
                chat_id=user_id,
                photo=open('pay/pay.jpeg', 'rb'),
                caption="""✨ Selamat! Anda telah disetujui untuk bergabung ke paket VIP berikut.

<blockquote>Sekarang kalian memiliki akses untuk bergabung keseluruhan tautan channel dan grup vvip di bawah. Silahkan gabung sekarang ke channel dan grup vvip yang kalian inginkan.</blockquote>
            """,
                reply_markup=reply_markup,
                parse_mode="HTML"
            )

            responding_admin_id = query.from_user.id 
            await query.answer(text="Pembayaran diterima. Pengguna sudah diundang ke grup.")

            # Kirim notifikasi ke admin
            for admin_id in ADMIN_CHAT_IDS:
                try:
                    if admin_id == responding_admin_id:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text="✅ Pembayaran diterima. Pengguna sudah diundang ke grup.",
                            reply_to_message_id=original_message_id
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=f"✅ Pembayaran diterima oleh admin {responding_admin_id}. Pengguna sudah diundang ke grup."
                        )
                except Exception as e:
                    print(f"Gagal mengirim pesan ke admin {admin_id}: {e}")

            # Kirim pesan sukses yang akan dihapus setelah 2 menit
            success_message = await context.bot.send_message(
                chat_id=user_id_from_data,
                text="✅ Selamat pembayaran anda diterima oleh admin..."
            )
            
            async def delete_message(context: ContextTypes.DEFAULT_TYPE):
                message = context.job.data
                try:
                    await message.delete()
                except Exception as e:
                    print(f"Error deleting message: {e}")

            context.job_queue.run_once(
                delete_message,
                120,
                data=success_message
            )

        except Exception as e:
            print(f"Error: {e}")
            await query.answer(text="Terjadi kesalahan dalam memproses pembayaran.")
        return


    # Jika admin menekan No
    elif action == "no":
        try:
            # Cek apakah file Excel sudah ada
            try:
                existing_data = pd.read_excel(file_name)
            except FileNotFoundError:
                # Jika file tidak ada, kirim pesan error dan return
                await context.bot.send_message(
                    chat_id=user_id,
                    text="Bukti pembayaran tidak valid mohon lakukan pembelian ulang."
                )
                await query.answer(text="Pembayaran ditolak.")
                return

        # Cek apakah user ada di database
            user_data = existing_data[existing_data['Chat ID'] == user_id]
        
        # Jika data pengguna tidak ditemukan
            if user_data.empty:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="❌ Bukti pembayaran tidak valid mohon lakukan pembelian ulang."
                )
            else:
                # Ambil index user
                user_index = user_data.index[0]
            
                # Ambil tanggal kadaluarsa saat ini
                current_expiry = datetime.strptime(existing_data.loc[user_index, 'Tanggal Kadaluarsa'], '%d %B %Y')
                current_date = datetime.now()
            
                # Hitung sisa hari saat ini
                current_remaining_days = (current_expiry - current_date).days
            
                # Jika sisa hari sudah 0 atau kurang
                if current_remaining_days <= 0:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="❌ Bukti pembayaran tidak valid mohon lakukan pembelian ulang."
                    )
                else:
                    # Kurangi 20 hari, tapi tidak kurang dari 0
                    new_remaining_days = max(0, current_remaining_days - 20)
                    new_expiry = current_date + timedelta(days=new_remaining_days)
                
                    # Update tanggal kadaluarsa
                    existing_data.loc[user_index, 'Tanggal Kadaluarsa'] = new_expiry.strftime('%d %B %Y')
                    existing_data.loc[user_index, 'Sisa Waktu'] = f"{new_remaining_days} hari"
                    existing_data.loc[user_index, 'Status VIP'] = "Tidak Aktif" if new_remaining_days <= 0 else "Aktif"
                
                    # Simpan perubahan ke file Excel
                    existing_data.to_excel(file_name, index=False)
                
                    if new_remaining_days <= 0:
                        await context.bot.send_message(
                            chat_id=user_id,
                            text="Bukti pembayaran tidak valid mohon lakukan pembelian ulang."
                    )
                    else:
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=f"❌ Maaf, bukti pembayaran Anda tidak valid. Waktu VIP Anda dikurangi 20 hari sebagai sanksi. Sisa waktu: {new_remaining_days} hari"
                        )
            
            responding_admin_id = query.from_user.id
            await query.answer(text="Pembayaran ditolak.")

            # Kirim notifikasi ke admin
            for admin_id in ADMIN_CHAT_IDS:
                try:
                    if admin_id == responding_admin_id:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text="❌ Pembayaran ditolak. Pengguna tidak ditambahkan ke grup.",
                            reply_to_message_id=original_message_id
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=f"❌ Pembayaran ditolak oleh admin {responding_admin_id}. Pengguna tidak ditambahkan ke grup."
                        )
                except Exception as e:
                    print(f"Gagal mengirim pesan ke admin {admin_id}: {e}")

        except Exception as e:
            print(f"Error: {e}")
            await query.answer(text="Terjadi kesalahan dalam memproses penolakan.")



# Fungsi untuk menghitung waktu berakhirnya VVIP
def cek_waktu_vvip_berakhir(waktu_pembelian):
    waktu_berakhir = waktu_pembelian + timedelta(days=30)  
    waktu_pemberitahuan = waktu_berakhir - timedelta(days=1)  # Pemberitahuan 24 jam sebelum berakhir
    return waktu_pemberitahuan, waktu_berakhir

# Fungsi untuk mengirimkan notifikasi
def kirim_notifikasi():
    print("VVIP akan segera berakhir, mohon beli kembali.")

# Fungsi utama untuk memeriksa dan memberi notifikasi
def periksa_vvip(waktu_pembelian):
    waktu_pemberitahuan, waktu_berakhir = cek_waktu_vvip_berakhir(waktu_pembelian)
    
    # Cek jika sekarang sudah 24 jam sebelum waktu berakhir
    if datetime.now() >= waktu_pemberitahuan and datetime.now() < waktu_berakhir:
        kirim_notifikasi()

# Contoh penggunaan
waktu_pembelian = datetime(2024, 12, 1, 10, 0) 
periksa_vvip(waktu_pembelian)


from telegram import Chat

async def check_expired_members():
    """
    Fungsi ini akan dijalankan setiap 24 jam untuk mengecek user dengan sisa waktu 0 hari
    dan mengeluarkannya dari grup jika ditemukan.
    """
    try:
        df = pd.read_excel(file_name)
        current_date = datetime.now()

        # Pastikan kolom 'Sisa Waktu' adalah string
        if 'Sisa Waktu' in df.columns:
            df['Sisa Waktu'] = df['Sisa Waktu'].astype(str)

        for index, row in df.iterrows():
            user_id = row['Chat ID']
            nama = row['Nama Lengkap']
            
            # Pastikan 'Sisa Waktu' adalah string dan tidak kosong
            sisa_waktu_str = row['Sisa Waktu']
            if isinstance(sisa_waktu_str, str):
                try:
                    sisa_hari = int(sisa_waktu_str.replace(" hari", "").strip())
                except ValueError:
                    sisa_hari = 0  # Atau nilai default lainnya jika parsing gagal
            else:
                sisa_hari = 0  # Atau nilai default lainnya jika bukan string

            # Jika waktu habis, kick user dari semua grup yang terdaftar
            if sisa_hari <= 0:
                for group_id in GROUP_IDS:
                    try:
                        # Mengecek tipe chat terlebih dahulu
                        chat = await bot.get_chat(group_id)
                        if chat.type in ["group", "supergroup"]:
                            await bot.ban_chat_member(chat_id=group_id, user_id=user_id)
                            await bot.send_message(
                                chat_id=group_id, 
                                text=f"⚠️ {nama} dengan ID {user_id} telah dikeluarkan dari grup karena masa berlangganan telah habis."
                            )
                            print(f"[INFO] {nama} dengan ID {user_id} berhasil dikeluarkan dari grup {group_id}")
                        elif chat.type == "channel":
                            await bot.ban_chat_member(chat_id=group_id, user_id=user_id)
                            print(f"[INFO] {nama} dengan ID {user_id} berhasil dihapus aksesnya dari channel {group_id}")
                    except Exception as e:
                        print(f"[ERROR] Gagal mengeluarkan {nama}: {e}")

                # Update status di file Excel
                df.at[index, 'Status VIP'] = "Tidak Aktif"
                df.at[index, 'Sisa Waktu'] = "0 hari"  # Tetap sebagai string

        # Simpan perubahan ke file
        df.to_excel(file_name, index=False)
        print("[INFO] Pemeriksaan selesai dan file diperbarui.")

    except FileNotFoundError:
        print("[ERROR] File Excel tidak ditemukan.")

async def schedule_daily_check():
    """Jadwal untuk melakukan pemeriksaan otomatis setiap 1 jam."""
    while True:
        update_excel_file(file_name)  # Update file Excel
        await check_expired_members()
        print(f"File telah diperbarui pada {datetime.now()}")
        await asyncio.sleep(3600)  # 3600 detik = 1 jam 


def main():
    # Menggunakan Application yang sesuai dengan async
    application = Application.builder().token(TOKEN).build()

    job_queue = application.job_queue
    loop = asyncio.get_event_loop()
    loop.create_task(schedule_daily_check())
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(beli_vvip, pattern='^beli_vvip$'))
    application.add_handler(CallbackQueryHandler(lambda u, c: adjust_month(u, c, -1), pattern='^kurang_bulan$'))
    application.add_handler(CallbackQueryHandler(lambda u, c: adjust_month(u, c, +1), pattern='^tambah_bulan$'))
    application.add_handler(CallbackQueryHandler(lanjut, pattern='^lanjut$'))
    application.add_handler(CallbackQueryHandler(batal, pattern='^batal$'))
    application.add_handler(CallbackQueryHandler(batal, pattern='^batalkan$'))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(CallbackQueryHandler(button))
    # application.add_handler(CallbackQueryHandler(handle_callback))

    print("Bot Telah DiJalankan")
    application.run_polling()

if __name__ == "__main__":
    main()
